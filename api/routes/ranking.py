"""
Blueprint de ranking por jugador.

Rutas:
  GET /api/ranking   → JSON con ranking acumulado por categoría
  GET /ranking       → Vista pública del ranking
"""

import logging
from collections import defaultdict

from io import BytesIO

from flask import Blueprint, jsonify, render_template, send_file
from openpyxl import Workbook



from config.settings import CATEGORIAS, COLORES_CATEGORIA, EMOJI_CATEGORIA

logger = logging.getLogger(__name__)

ranking_bp = Blueprint('ranking', __name__)


def _sb():
    from utils.supabase_client import get_supabase_admin
    return get_supabase_admin()


def _use_supabase() -> bool:
    from config.settings import SUPABASE_URL, SUPABASE_SERVICE_ROLE_KEY
    return bool(SUPABASE_URL and SUPABASE_SERVICE_ROLE_KEY)


def _calcular_ranking() -> dict:
    """Devuelve ranking agrupado por categoría (todos los torneos finalizados).

    Estructura de retorno:
    {
      "Tercera": [
        {"posicion": 1, "nombre": "Juan", "apellido": "García",
         "jugador_id": "uuid", "puntos": 250, "torneos": 2,
         "mejor_resultado": "campeon"},
        ...
      ],
      ...
    }
    """
    if not _use_supabase():
        return {}

    sb = _sb()

    # 1. Todos los torneos finalizados
    torneos_resp = (
        sb.table('torneos')
        .select('id')
        .eq('estado', 'finalizado')
        .execute()
    )
    torneos_finalizados = torneos_resp.data or []
    if not torneos_finalizados:
        return {}

    torneo_ids = [t['id'] for t in torneos_finalizados]

    # 2. Puntos de esos torneos
    puntos_resp = (
        sb.table('puntos_jugador')
        .select('jugador_id, torneo_id, categoria, puntos, concepto')
        .in_('torneo_id', torneo_ids)
        .execute()
    )
    filas = puntos_resp.data or []
    if not filas:
        return {}

    # 3. Jugadores involucrados
    jugador_ids = list({f['jugador_id'] for f in filas})
    jugadores_resp = (
        sb.table('jugadores')
        .select('id, nombre, apellido')
        .in_('id', jugador_ids)
        .execute()
    )
    jugadores_map = {j['id']: j for j in (jugadores_resp.data or [])}

    # 4. Agregar en Python: sum(puntos), count(torneos), mejor concepto por jugador+categoría
    _ORDEN_CONCEPTO = ['serie', 'octavos', 'cuartos', 'semifinal', 'vicecampeon', 'campeon']

    acum: dict[str, dict] = {}  # key: "jugador_id|categoria"
    for fila in filas:
        key = f"{fila['jugador_id']}|{fila['categoria']}"
        if key not in acum:
            acum[key] = {
                'jugador_id':      fila['jugador_id'],
                'categoria':       fila['categoria'],
                'puntos':          0,
                'torneos':         0,
                'mejor_resultado': 'serie',
            }
        acum[key]['puntos'] += fila['puntos']
        acum[key]['torneos'] += 1
        concepto_actual = acum[key]['mejor_resultado']
        concepto_nuevo = fila['concepto']
        if _ORDEN_CONCEPTO.index(concepto_nuevo) > _ORDEN_CONCEPTO.index(concepto_actual):
            acum[key]['mejor_resultado'] = concepto_nuevo

    # 5. Agrupar por categoría y ordenar
    ranking: dict[str, list] = defaultdict(list)
    for entry in acum.values():
        jugador = jugadores_map.get(entry['jugador_id'], {})
        ranking[entry['categoria']].append({
            'jugador_id':      entry['jugador_id'],
            'nombre':          jugador.get('nombre', ''),
            'apellido':        jugador.get('apellido', ''),
            'puntos':          entry['puntos'],
            'torneos':         entry['torneos'],
            'mejor_resultado': entry['mejor_resultado'],
        })

    for cat in ranking:
        ranking[cat].sort(key=lambda x: -x['puntos'])
        for i, row in enumerate(ranking[cat], start=1):
            row['posicion'] = i

    # Respetar orden canónico de categorías
    return {cat: ranking[cat] for cat in CATEGORIAS if cat in ranking}


def _build_ranking_workbook(ranking: dict) -> BytesIO:
    """Construye un workbook Excel a partir del dict de ranking.

    - Una hoja por categoría con columnas: Posición, Nombre, Apellido, Puntos, Torneos, Mejor Resultado.
    - Si ranking está vacío, genera una hoja 'Sin datos' con mensaje informativo.
    - Retorna BytesIO listo para ser consumido por send_file.
    """
    wb = Workbook()
    wb.remove(wb.active)  # eliminar hoja vacía por defecto

    if not ranking:
        ws = wb.create_sheet('Sin datos')
        ws.append(['No hay torneos finalizados'])
    else:
        headers = ['Posición', 'Nombre', 'Apellido', 'Puntos', 'Torneos', 'Mejor Resultado']
        for cat, entries in ranking.items():
            ws = wb.create_sheet(cat[:31])
            ws.append(headers)
            for e in entries:
                ws.append([
                    e['posicion'],
                    e['nombre'],
                    e['apellido'],
                    e['puntos'],
                    e['torneos'],
                    e['mejor_resultado'],
                ])

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


@ranking_bp.route('/api/ranking')
def api_ranking():
    try:
        data = _calcular_ranking()
        return jsonify({'ranking': data})
    except Exception:
        logger.exception('Error al calcular ranking')
        return jsonify({'error': 'Error al calcular el ranking'}), 500


@ranking_bp.route('/ranking')
def ranking_publico():
    try:
        ranking = _calcular_ranking()
    except Exception:
        logger.exception('Error al renderizar ranking')
        ranking = {}

    return render_template(
        'ranking.html',
        ranking=ranking,
        colores=COLORES_CATEGORIA,
        emojis=EMOJI_CATEGORIA,
    )


@ranking_bp.route('/api/admin/ranking/export')
def exportar_ranking():
    """Exporta el ranking acumulado como archivo Excel (.xlsx).

    Sólo accesible para administradores — protegido por el middleware
    before_request de main.py (prefijo /api/admin/ NO está en la whitelist).
    """
    try:
        ranking = _calcular_ranking()
    except Exception as exc:
        logger.exception('Error al exportar ranking')
        return jsonify({'error': str(exc)}), 500

    buf = _build_ranking_workbook(ranking)
    return send_file(
        buf,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='ranking.xlsx',
    )
