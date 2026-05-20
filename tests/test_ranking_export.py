"""Tests para la exportación Excel del ranking (F3).

T1: _build_ranking_workbook con datos y vacío (RED primero, sin implementación).
T2: GET /api/admin/ranking/export — rutas, seguridad, error handling.
"""

import pytest
from io import BytesIO
from unittest.mock import patch

import openpyxl

# ── Fixture compartida ────────────────────────────────────────────────────────

RANKING_FIXTURE = {
    "Cuarta": [
        {"posicion": 1, "jugador_id": "j1", "nombre": "Ana", "apellido": "García",
         "puntos": 200, "torneos": 2, "mejor_resultado": "campeon"},
        {"posicion": 2, "jugador_id": "j2", "nombre": "Luis", "apellido": "Pérez",
         "puntos": 100, "torneos": 1, "mejor_resultado": "serie"},
    ],
    "Tercera": [
        {"posicion": 1, "jugador_id": "j3", "nombre": "Roberto", "apellido": "Silva",
         "puntos": 150, "torneos": 1, "mejor_resultado": "semifinal"},
    ],
}

EXPECTED_HEADERS = ["Posición", "Nombre", "Apellido", "Puntos", "Torneos", "Mejor Resultado"]


# ── T1: _build_ranking_workbook ───────────────────────────────────────────────

class TestBuildRankingWorkbook:
    def test_build_workbook_with_data(self):
        """Workbook tiene una hoja por categoría, headers correctos y primera fila de datos."""
        from api.routes.ranking import _build_ranking_workbook

        buf = _build_ranking_workbook(RANKING_FIXTURE)

        assert isinstance(buf, BytesIO)
        wb = openpyxl.load_workbook(buf)

        # Una hoja por categoría
        assert set(wb.sheetnames) == {"Cuarta", "Tercera"}

        # Header row en hoja Cuarta
        ws_cuarta = wb["Cuarta"]
        header = [cell.value for cell in ws_cuarta[1]]
        assert header == EXPECTED_HEADERS

        # Primera fila de datos coincide con el fixture
        primera_fila = [cell.value for cell in ws_cuarta[2]]
        assert primera_fila[0] == 1          # posicion
        assert primera_fila[1] == "Ana"      # nombre
        assert primera_fila[2] == "García"   # apellido
        assert primera_fila[3] == 200        # puntos
        assert primera_fila[4] == 2          # torneos
        assert primera_fila[5] == "campeon"  # mejor_resultado

    def test_build_workbook_empty(self):
        """Ranking vacío → hoja única 'Sin datos' con mensaje informativo."""
        from api.routes.ranking import _build_ranking_workbook

        buf = _build_ranking_workbook({})

        assert isinstance(buf, BytesIO)
        wb = openpyxl.load_workbook(buf)

        assert len(wb.sheetnames) == 1
        assert wb.sheetnames[0] == "Sin datos"

        # Tiene al menos una fila con contenido
        ws = wb.active
        first_row = [cell.value for cell in ws[1] if cell.value is not None]
        assert len(first_row) > 0


# ── T2: GET /api/admin/ranking/export ─────────────────────────────────────────

XLSX_MIMETYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


class TestExportRankingRoute:
    def test_export_route_returns_xlsx(self, admin_cookie):
        """Admin + datos → 200 con Content-Type xlsx y body no vacío."""
        with patch("api.routes.ranking._calcular_ranking", return_value=RANKING_FIXTURE):
            resp = admin_cookie.get("/api/admin/ranking/export")

        assert resp.status_code == 200
        assert XLSX_MIMETYPE in resp.content_type
        assert len(resp.data) > 0

    def test_export_route_security(self, client):
        """Sin cookie de admin → 302 redirect a /login (ADR-001 — más importante)."""
        resp = client.get("/api/admin/ranking/export")
        assert resp.status_code == 302
        assert "/login" in resp.headers.get("Location", "")

    def test_export_route_empty_ranking(self, admin_cookie):
        """Ranking vacío → sigue siendo 200 + xlsx válido (no error)."""
        with patch("api.routes.ranking._calcular_ranking", return_value={}):
            resp = admin_cookie.get("/api/admin/ranking/export")

        assert resp.status_code == 200
        assert XLSX_MIMETYPE in resp.content_type
        assert len(resp.data) > 0
        # Debe ser un xlsx válido (no corrompido)
        wb = openpyxl.load_workbook(BytesIO(resp.data))
        assert len(wb.sheetnames) == 1
        assert wb.sheetnames[0] == "Sin datos"

    def test_export_route_calcular_raises(self, admin_cookie):
        """_calcular_ranking lanza excepción → 500 JSON con clave 'error'."""
        with patch("api.routes.ranking._calcular_ranking", side_effect=Exception("db error")):
            resp = admin_cookie.get("/api/admin/ranking/export")

        assert resp.status_code == 500
        assert "error" in resp.get_json()
